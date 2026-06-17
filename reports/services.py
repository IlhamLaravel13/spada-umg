import io
import csv
import logging
from datetime import datetime
from django.db.models import Count, Q, Avg, Sum
from django.db import transaction
from django.utils import timezone
from django.template.loader import render_to_string
from .repositories import ReportRepository
from .models import Report

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, Image
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


class ReportService:
    def __init__(self):
        self.repo = ReportRepository()

    def generate_report(self, report_type: str, fmt: str, title: str,
                        parameters: dict, user) -> dict:
        title = title or dict(Report.TYPE_CHOICES).get(report_type, 'Report')
        report = self.repo.create_report(title, report_type, fmt, parameters, user)
        try:
            with transaction.atomic():
                if report_type == 'academic':
                    data = self._get_academic_data(parameters)
                elif report_type == 'attendance':
                    data = self._get_attendance_data(parameters)
                elif report_type == 'grade':
                    data = self._get_grade_data(parameters)
                elif report_type == 'payment':
                    data = self._get_payment_data(parameters)
                elif report_type == 'user':
                    data = self._get_user_data(parameters)
                elif report_type == 'course':
                    data = self._get_course_data(parameters)
                else:
                    return {'success': False, 'error': 'Invalid report type'}

                file_path = self._build_file(report, fmt, data)
                self.repo.mark_ready(report.id, file_path)
                return {'success': True, 'report': report}
        except Exception as e:
            logger.exception(f"Report generation failed: {e}")
            return {'success': False, 'error': str(e)}

    def _get_academic_data(self, params):
        from academics.models import Faculty, StudyProgram, Course, Class, Enrollment, Semester, AcademicYear
        data = {
            'title': 'Academic Report',
            'generated_at': timezone.now(),
            'faculties': [],
        }
        faculties = Faculty.objects.filter(is_active=True)
        if params.get('faculty_id'):
            faculties = faculties.filter(id=params['faculty_id'])
        for faculty in faculties:
            prog_count = faculty.study_programs.filter(is_active=True).count()
            course_count = Course.objects.filter(
                study_program__faculty=faculty, is_active=True
            ).count()
            student_count = Enrollment.objects.filter(
                class_enrolled__course__study_program__faculty=faculty,
                status='active'
            ).values('student').distinct().count()
            data['faculties'].append({
                'name': faculty.name,
                'code': faculty.code,
                'study_programs': prog_count,
                'courses': course_count,
                'students': student_count,
            })
        data['total_faculties'] = faculties.count()
        data['total_study_programs'] = StudyProgram.objects.filter(is_active=True).count()
        data['total_courses'] = Course.objects.filter(is_active=True).count()
        data['total_enrollments'] = Enrollment.objects.filter(status='active').count()
        return data

    def _get_attendance_data(self, params):
        from academics.models import Class, Enrollment
        from attendance.models import Attendance, AttendanceSession
        data = {
            'title': 'Attendance Report',
            'generated_at': timezone.now(),
            'classes': [],
        }
        sessions = AttendanceSession.objects.all()
        if params.get('class_id'):
            sessions = sessions.filter(class_meta_id=params['class_id'])
        if params.get('start_date'):
            sessions = sessions.filter(date__gte=params['start_date'])
        if params.get('end_date'):
            sessions = sessions.filter(date__lte=params['end_date'])
        if params.get('semester_id'):
            sessions = sessions.filter(class_meta__semester_id=params['semester_id'])

        for session in sessions.select_related('class_meta', 'class_meta__course').prefetch_related('attendances')[:100]:
            total = session.attendances.count()
            present = session.attendances.filter(status='present').count()
            late = session.attendances.filter(status='late').count()
            absent = session.attendances.filter(status='absent').count()
            excused = session.attendances.filter(status='excused').count()
            data['classes'].append({
                'class_name': str(session.class_meta),
                'date': session.date,
                'topic': session.topic,
                'total': total,
                'present': present,
                'late': late,
                'absent': absent,
                'excused': excused,
                'attendance_rate': round((present + late) / total * 100, 1) if total else 0,
            })
        data['total_sessions'] = len(data['classes'])
        return data

    def _get_grade_data(self, params):
        from academics.models import Class, Enrollment
        data = {
            'title': 'Grade Report',
            'generated_at': timezone.now(),
            'classes': [],
        }
        enrollments = Enrollment.objects.filter(status__in=['active', 'completed'])
        if params.get('class_id'):
            enrollments = enrollments.filter(class_enrolled_id=params['class_id'])
        if params.get('course_id'):
            enrollments = enrollments.filter(class_enrolled__course_id=params['course_id'])
        if params.get('semester_id'):
            enrollments = enrollments.filter(class_enrolled__semester_id=params['semester_id'])
        if params.get('study_program_id'):
            enrollments = enrollments.filter(
                class_enrolled__course__study_program_id=params['study_program_id']
            )

        enrollments = enrollments.select_related(
            'student', 'class_enrolled', 'class_enrolled__course'
        ).order_by('class_enrolled__course__code')[:200]

        grades_by_class = {}
        for e in enrollments:
            key = str(e.class_enrolled)
            if key not in grades_by_class:
                grades_by_class[key] = {
                    'class_name': key,
                    'course_code': e.class_enrolled.course.code,
                    'course_name': e.class_enrolled.course.name,
                    'students': [],
                    'total': 0,
                    'sum': 0,
                }
            grades_by_class[key]['students'].append({
                'student_name': e.student.get_full_name() or e.student.username,
                'nim': e.student.nim,
                'grade': float(e.grade_final) if e.grade_final else None,
                'grade_letter': e.grade_letter,
                'status': e.status,
            })
            if e.grade_final:
                grades_by_class[key]['total'] += 1
                grades_by_class[key]['sum'] += float(e.grade_final)

        for k, v in grades_by_class.items():
            v['average'] = round(v['sum'] / v['total'], 2) if v['total'] else 0
            data['classes'].append(v)

        data['total_classes'] = len(data['classes'])
        return data

    def _get_payment_data(self, params):
        from payments.models import Payment
        payments = Payment.objects.all()
        if params.get('start_date'):
            payments = payments.filter(created_at__gte=params['start_date'])
        if params.get('end_date'):
            payments = payments.filter(created_at__lte=params['end_date'])
        if params.get('status'):
            payments = payments.filter(status=params['status'])
        payments = payments.select_related('user').order_by('-created_at')[:200]
        data = {
            'title': 'Payment Report',
            'generated_at': timezone.now(),
            'payments': [
                {
                    'invoice': p.invoice_number,
                    'user': p.user.get_full_name() or p.user.username,
                    'type': p.get_payment_type_display(),
                    'amount': float(p.amount),
                    'status': p.get_status_display(),
                    'method': p.get_payment_method_display(),
                    'date': p.created_at,
                }
                for p in payments
            ],
            'total_amount': sum(float(p.amount) for p in payments),
            'total_payments': payments.count(),
        }
        return data

    def _get_user_data(self, params):
        from accounts.models import User
        users = User.objects.all()
        if params.get('role'):
            users = users.filter(role=params['role'])
        if params.get('study_program_id'):
            users = users.filter(study_program_id=params['study_program_id'])
        if params.get('faculty_id'):
            users = users.filter(faculty_id=params['faculty_id'])
        data = {
            'title': 'User Report',
            'generated_at': timezone.now(),
            'users': [
                {
                    'username': u.username,
                    'name': u.get_full_name() or u.username,
                    'email': u.email,
                    'role': u.get_role_display(),
                    'nim': u.nim,
                    'nidn': u.nidn,
                    'is_verified': u.is_verified,
                    'is_active': u.is_active,
                    'date_joined': u.date_joined,
                }
                for u in users.select_related('faculty', 'study_program')[:200]
            ],
            'total_users': users.count(),
            'active_users': users.filter(is_active=True).count(),
        }
        return data

    def _get_course_data(self, params):
        from academics.models import Course, Class, Enrollment
        courses = Course.objects.filter(is_active=True)
        if params.get('faculty_id'):
            courses = courses.filter(study_program__faculty_id=params['faculty_id'])
        if params.get('study_program_id'):
            courses = courses.filter(study_program_id=params['study_program_id'])
        data = {
            'title': 'Course Report',
            'generated_at': timezone.now(),
            'courses': [],
        }
        for course in courses.select_related('study_program', 'study_program__faculty').prefetch_related('classes', 'classes__enrollments')[:100]:
            total_classes = course.classes.filter(is_active=True).count()
            total_enrollments = Enrollment.objects.filter(
                class_enrolled__course=course, status='active'
            ).count()
            data['courses'].append({
                'code': course.code,
                'name': course.name,
                'study_program': course.study_program.name,
                'faculty': course.study_program.faculty.name,
                'credits': course.credits,
                'classes': total_classes,
                'enrollments': total_enrollments,
            })
        data['total_courses'] = len(data['courses'])
        return data

    def _build_file(self, report: Report, fmt: str, data: dict) -> str:
        if fmt == 'pdf':
            return self._generate_pdf(report, data)
        elif fmt == 'excel':
            return self._generate_excel(report, data)
        elif fmt == 'csv':
            return self._generate_csv(report, data)
        return ''

    def _generate_pdf(self, report: Report, data: dict) -> str:
        if not HAS_REPORTLAB:
            logger.warning("ReportLab not installed, using fallback")
            return self._generate_csv(report, data)
        from django.conf import settings
        import os
        from django.core.files.base import ContentFile
        from django.core.files.storage import default_storage

        filename = f"reports/{report.report_type}_{report.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            rightMargin=30*mm, leftMargin=30*mm,
            topMargin=20*mm, bottomMargin=20*mm,
        )
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Title'],
            fontSize=18, spaceAfter=20,
            textColor=colors.HexColor('#1a365d'),
        )
        heading_style = ParagraphStyle(
            'CustomHeading', parent=styles['Heading2'],
            fontSize=12, spaceAfter=10,
            textColor=colors.HexColor('#2d3748'),
        )
        normal_style = ParagraphStyle(
            'CustomNormal', parent=styles['Normal'],
            fontSize=9, spaceAfter=6,
        )

        elements.append(Paragraph(f"SPADA UMG - {data.get('title', 'Report')}", title_style))
        elements.append(Paragraph(
            f"Generated: {data.get('generated_at', timezone.now()).strftime('%d %B %Y %H:%M')}",
            normal_style
        ))
        elements.append(Spacer(1, 12))

        if report.report_type == 'academic':
            elements = self._pdf_academic(elements, data, heading_style, normal_style)
        elif report.report_type == 'attendance':
            elements = self._pdf_attendance(elements, data, heading_style, normal_style)
        elif report.report_type == 'grade':
            elements = self._pdf_grade(elements, data, heading_style, normal_style)
        elif report.report_type == 'payment':
            elements = self._pdf_payment(elements, data, heading_style, normal_style)
        elif report.report_type == 'user':
            elements = self._pdf_user(elements, data, heading_style, normal_style)
        elif report.report_type == 'course':
            elements = self._pdf_course(elements, data, heading_style, normal_style)

        elements.append(Spacer(1, 20))
        elements.append(Paragraph(
            f"Report #{report.id} | SPADA UMG - Universitas Muhammadiyah Gresik",
            ParagraphStyle('Footer', parent=normal_style, fontSize=7, textColor=colors.grey)
        ))

        doc.build(elements)
        buffer.seek(0)
        saved_path = default_storage.save(filename, ContentFile(buffer.read()))
        buffer.close()
        return saved_path

    def _make_table(self, headers, rows, col_widths=None):
        table_data = [headers] + rows
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        header_fill = PatternFill if False else None
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FAFC')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ])
        t.setStyle(style)
        return t

    def _pdf_academic(self, elements, data, heading_style, normal_style):
        elements.append(Paragraph("Academic Report Summary", heading_style))
        summary_data = [
            ['Metric', 'Count'],
            ['Faculties', str(data.get('total_faculties', 0))],
            ['Study Programs', str(data.get('total_study_programs', 0))],
            ['Courses', str(data.get('total_courses', 0))],
            ['Active Enrollments', str(data.get('total_enrollments', 0))],
        ]
        elements.append(self._make_table(summary_data[0], summary_data[1:], col_widths=[200, 100]))
        elements.append(Spacer(1, 12))

        if data.get('faculties'):
            elements.append(Paragraph("Faculty Details", heading_style))
            fac_headers = ['Faculty', 'Code', 'Study Programs', 'Courses', 'Students']
            fac_rows = [
                [f['name'], f['code'], str(f['study_programs']), str(f['courses']), str(f['students'])]
                for f in data['faculties']
            ]
            elements.append(self._make_table(fac_headers, fac_rows))
        return elements

    def _pdf_attendance(self, elements, data, heading_style, normal_style):
        elements.append(Paragraph("Attendance Report", heading_style))
        if data.get('classes'):
            att_headers = ['Class', 'Date', 'Topic', 'Total', 'Present', 'Late', 'Absent', 'Rate %']
            att_rows = [
                [c['class_name'], str(c['date']), c['topic'][:30],
                 str(c['total']), str(c['present']), str(c['late']),
                 str(c['absent']), str(c['attendance_rate'])]
                for c in data['classes']
            ]
            elements.append(self._make_table(att_headers, att_rows))
            elements.append(Spacer(1, 8))
            total_sessions = data.get('total_sessions', 0)
            total_all = sum(c['total'] for c in data['classes'])
            present_all = sum(c['present'] for c in data['classes'])
            rate = round(present_all / total_all * 100, 1) if total_all else 0
            elements.append(Paragraph(
                f"Total Sessions: {total_sessions} | Overall Attendance Rate: {rate}%",
                normal_style
            ))
        return elements

    def _pdf_grade(self, elements, data, heading_style, normal_style):
        elements.append(Paragraph("Grade Report", heading_style))
        if data.get('classes'):
            for cls in data['classes']:
                elements.append(Paragraph(
                    f"{cls['class_name']} - Avg: {cls['average']}",
                    heading_style
                ))
                grade_headers = ['Student', 'NIM', 'Grade', 'Letter', 'Status']
                grade_rows = [
                    [s['student_name'], s['nim'] or '-',
                     str(s['grade']) if s['grade'] else '-',
                     s['grade_letter'] or '-', s['status']]
                    for s in cls['students']
                ]
                elements.append(self._make_table(grade_headers, grade_rows))
                elements.append(Spacer(1, 8))
        return elements

    def _pdf_payment(self, elements, data, heading_style, normal_style):
        elements.append(Paragraph("Payment Report", heading_style))
        elements.append(Paragraph(
            f"Total Payments: {data.get('total_payments', 0)} | "
            f"Total Amount: Rp {data.get('total_amount', 0):,.2f}",
            normal_style
        ))
        elements.append(Spacer(1, 8))
        if data.get('payments'):
            pay_headers = ['Invoice', 'User', 'Type', 'Amount', 'Status', 'Date']
            pay_rows = [
                [p['invoice'], p['user'], p['type'],
                 f"Rp {p['amount']:,.0f}", p['status'],
                 p['date'].strftime('%Y-%m-%d') if hasattr(p['date'], 'strftime') else str(p['date'])]
                for p in data['payments']
            ]
            elements.append(self._make_table(pay_headers, pay_rows))
        return elements

    def _pdf_user(self, elements, data, heading_style, normal_style):
        elements.append(Paragraph("User Report", heading_style))
        elements.append(Paragraph(
            f"Total Users: {data.get('total_users', 0)} | "
            f"Active: {data.get('active_users', 0)}",
            normal_style
        ))
        elements.append(Spacer(1, 8))
        if data.get('users'):
            user_headers = ['Username', 'Name', 'Email', 'Role', 'NIM/NIDN', 'Active']
            user_rows = [
                [u['username'], u['name'], u['email'], u['role'],
                 u['nim'] or u['nidn'] or '-', 'Yes' if u['is_active'] else 'No']
                for u in data['users']
            ]
            elements.append(self._make_table(user_headers, user_rows))
        return elements

    def _pdf_course(self, elements, data, heading_style, normal_style):
        elements.append(Paragraph("Course Report", heading_style))
        elements.append(Paragraph(
            f"Total Courses: {data.get('total_courses', 0)}",
            normal_style
        ))
        elements.append(Spacer(1, 8))
        if data.get('courses'):
            course_headers = ['Code', 'Name', 'Study Program', 'Faculty', 'Credits', 'Classes', 'Students']
            course_rows = [
                [c['code'], c['name'][:35], c['study_program'], c['faculty'],
                 str(c['credits']), str(c['classes']), str(c['enrollments'])]
                for c in data['courses']
            ]
            elements.append(self._make_table(course_headers, course_rows))
        return elements

    def _generate_excel(self, report: Report, data: dict) -> str:
        if not HAS_OPENPYXL:
            logger.warning("openpyxl not installed, using CSV fallback")
            return self._generate_csv(report, data)
        from django.conf import settings
        import os
        from django.core.files.base import ContentFile
        from django.core.files.storage import default_storage

        filename = f"reports/{report.report_type}_{report.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = data.get('title', 'Report')

        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        if report.report_type == 'academic' and data.get('faculties'):
            headers = ['Faculty', 'Code', 'Study Programs', 'Courses', 'Students']
            ws.append(['ACADEMIC REPORT', '', '', '', ''])
            ws.merge_cells('A1:E1')
            ws.append([])
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            for f in data['faculties']:
                ws.append([f['name'], f['code'], f['study_programs'], f['courses'], f['students']])
            ws.append([])
            ws.append(['Summary', '', '', '', ''])
            ws.append(['Total Faculties', data.get('total_faculties', 0), '', '', ''])
            ws.append(['Total Study Programs', data.get('total_study_programs', 0), '', '', ''])
            ws.append(['Total Courses', data.get('total_courses', 0), '', '', ''])
            ws.append(['Total Enrollments', data.get('total_enrollments', 0), '', '', ''])
        elif report.report_type == 'attendance' and data.get('classes'):
            headers = ['Class', 'Date', 'Topic', 'Total', 'Present', 'Late', 'Absent', 'Excused', 'Rate %']
            ws.append(['ATTENDANCE REPORT', '', '', '', '', '', '', '', ''])
            ws.merge_cells('A1:I1')
            ws.append([])
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
            for c in data['classes']:
                ws.append([c['class_name'], str(c['date']), c['topic'],
                          c['total'], c['present'], c['late'], c['absent'],
                          c['excused'], c['attendance_rate']])
            ws.append([])
            ws.append(['Total Sessions', data.get('total_sessions', 0), '', '', '', '', '', '', ''])
        elif report.report_type == 'grade' and data.get('classes'):
            ws.append(['GRADE REPORT', '', '', '', ''])
            ws.merge_cells('A1:E1')
            ws.append([])
            row_num = 3
            for cls in data['classes']:
                ws.cell(row=row_num, column=1, value=f"Class: {cls['class_name']}")
                ws.cell(row=row_num, column=1).font = Font(bold=True, size=10)
                row_num += 1
                ws.cell(row=row_num, column=1, value=f"Average Grade: {cls['average']}")
                row_num += 1
                headers = ['Student', 'NIM', 'Grade', 'Grade Letter', 'Status']
                ws.append(headers)
                row_num += 1
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
        elif report.report_type == 'payment' and data.get('payments'):
            headers = ['Invoice', 'User', 'Type', 'Amount', 'Status', 'Method', 'Date']
            ws.append(['PAYMENT REPORT', '', '', '', '', '', ''])
            ws.merge_cells('A1:G1')
            ws.append([])
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
            for p in data['payments']:
                ws.append([p['invoice'], p['user'], p['type'], p['amount'],
                          p['status'], p['method'],
                          p['date'].strftime('%Y-%m-%d') if hasattr(p['date'], 'strftime') else str(p['date'])])
        elif report.report_type == 'user' and data.get('users'):
            headers = ['Username', 'Name', 'Email', 'Role', 'NIM/NIDN', 'Verified', 'Active']
            ws.append(['USER REPORT', '', '', '', '', '', ''])
            ws.merge_cells('A1:G1')
            ws.append([])
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
            for u in data['users']:
                ws.append([u['username'], u['name'], u['email'], u['role'],
                          u['nim'] or u['nidn'] or '-',
                          'Yes' if u['is_verified'] else 'No',
                          'Yes' if u['is_active'] else 'No'])
        elif report.report_type == 'course' and data.get('courses'):
            headers = ['Code', 'Name', 'Study Program', 'Faculty', 'Credits', 'Classes', 'Students']
            ws.append(['COURSE REPORT', '', '', '', '', '', ''])
            ws.merge_cells('A1:G1')
            ws.append([])
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
            for c in data['courses']:
                ws.append([c['code'], c['name'], c['study_program'], c['faculty'],
                          c['credits'], c['classes'], c['enrollments']])

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        saved_path = default_storage.save(filename, ContentFile(buffer.read()))
        buffer.close()
        return saved_path

    def _generate_csv(self, report: Report, data: dict) -> str:
        from django.core.files.base import ContentFile
        from django.core.files.storage import default_storage

        filename = f"reports/{report.report_type}_{report.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        writer.writerow([data.get('title', 'Report')])
        writer.writerow(['Generated:', data.get('generated_at', timezone.now()).strftime('%Y-%m-%d %H:%M')])
        writer.writerow([])

        if report.report_type == 'academic' and data.get('faculties'):
            writer.writerow(['Faculty', 'Code', 'Study Programs', 'Courses', 'Students'])
            for f in data['faculties']:
                writer.writerow([f['name'], f['code'], f['study_programs'], f['courses'], f['students']])
            writer.writerow([])
            writer.writerow(['Total Faculties:', data.get('total_faculties', 0)])
            writer.writerow(['Total Study Programs:', data.get('total_study_programs', 0)])
            writer.writerow(['Total Courses:', data.get('total_courses', 0)])
            writer.writerow(['Total Enrollments:', data.get('total_enrollments', 0)])
        elif report.report_type == 'attendance' and data.get('classes'):
            writer.writerow(['Class', 'Date', 'Topic', 'Total', 'Present', 'Late', 'Absent', 'Excused', 'Rate'])
            for c in data['classes']:
                writer.writerow([c['class_name'], c['date'], c['topic'],
                                c['total'], c['present'], c['late'],
                                c['absent'], c['excused'], c['attendance_rate']])
        elif report.report_type == 'grade' and data.get('classes'):
            writer.writerow(['Class', 'Student', 'NIM', 'Grade', 'Grade Letter', 'Status'])
            for cls in data['classes']:
                for s in cls['students']:
                    writer.writerow([cls['class_name'], s['student_name'], s['nim'],
                                    s['grade'], s['grade_letter'], s['status']])
        elif report.report_type == 'payment' and data.get('payments'):
            writer.writerow(['Invoice', 'User', 'Type', 'Amount', 'Status', 'Method', 'Date'])
            for p in data['payments']:
                writer.writerow([p['invoice'], p['user'], p['type'], p['amount'],
                                p['status'], p['method'],
                                p['date'].strftime('%Y-%m-%d') if hasattr(p['date'], 'strftime') else str(p['date'])])
        elif report.report_type == 'user' and data.get('users'):
            writer.writerow(['Username', 'Name', 'Email', 'Role', 'NIM/NIDN', 'Verified', 'Active'])
            for u in data['users']:
                writer.writerow([u['username'], u['name'], u['email'], u['role'],
                                u['nim'] or u['nidn'] or '-', u['is_verified'], u['is_active']])
        elif report.report_type == 'course' and data.get('courses'):
            writer.writerow(['Code', 'Name', 'Study Program', 'Faculty', 'Credits', 'Classes', 'Students'])
            for c in data['courses']:
                writer.writerow([c['code'], c['name'], c['study_program'], c['faculty'],
                                c['credits'], c['classes'], c['enrollments']])

        buffer.seek(0)
        saved_path = default_storage.save(filename, ContentFile(buffer.getvalue().encode('utf-8-sig')))
        buffer.close()
        return saved_path

    def get_reports(self, user=None):
        if user and not user.is_anonymous and not user.is_admin():
            return self.repo.get_by_user(user.id)
        return self.repo.get_all()

    def get_report(self, report_id: int):
        return self.repo.get_by_id(report_id)

    def delete_report(self, report_id: int, user) -> dict:
        report = self.repo.get_by_id(report_id)
        if not report:
            return {'success': False, 'error': 'Report not found'}
        if report.generated_by != user and not user.is_admin():
            return {'success': False, 'error': 'Permission denied'}
        self.repo.delete_report(report_id)
        return {'success': True}
